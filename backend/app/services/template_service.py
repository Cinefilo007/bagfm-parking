import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment

class TemplateService:
    def generar_excel_socios_template(self) -> bytes:
        """
        Genera un archivo Excel con los encabezados necesarios para importar socios.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IMPORTAR_SOCIOS"

        # Encabezados
        headers = [
            "CEDULA", "NOMBRE", "APELLIDO", "EMAIL", "TELEFONO", 
            "PLACA", "MARCA", "MODELO", "COLOR"
        ]
        
        # Estilos para el encabezado
        header_fill = PatternFill(start_color="4EDEA3", end_color="4EDEA3", fill_type="solid")
        header_font = Font(bold=True, color="003824")
        center_align = Alignment(horizontal="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            # Ajustar ancho de columna
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

        # Instrucciones de ejemplo en la fila 2 (opcional, pero ayuda)
        ejemplo = [
            "V12345678", "JUAN", "PEREZ", "juan@ejemplo.com", "04121234567",
            "ABC12D", "TOYOTA", "COROLLA", "BLANCO"
        ]
        for col_num, value in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def generar_excel_pases_template(self) -> bytes:
        """
        Genera un archivo Excel con los encabezados necesarios para pases con identificación (Tipo B).
        Estructura Aegis Tactical v2.2 (20 Columnas): 
        Datos Personales (4) + 4 Vehículos (Marca, Modelo, Color, Placa = 16 columnas).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DATOS_IDENTIFICACION"

        # Encabezados tácticos v2.2 (20 Columnas)
        headers = [
            "NOMBRE COMPLETO", "CEDULA", "EMAIL", "TELEFONO", 
            "V1_PLACA", "V1_MARCA", "V1_MODELO", "V1_COLOR",
            "V2_PLACA", "V2_MARCA", "V2_MODELO", "V2_COLOR",
            "V3_PLACA", "V3_MARCA", "V3_MODELO", "V3_COLOR",
            "V4_PLACA", "V4_MARCA", "V4_MODELO", "V4_COLOR"
        ]
        
        # Estilos
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark tactico
        v_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Gris tactico para vehiculos
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_align
            if col_num <= 4:
                cell.fill = header_fill
            else:
                cell.fill = v_fill
                
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        # Ejemplo táctico
        ejemplo = [
            "PEDRO PEREZ", "V12345678", "pedro@aegis.com", "04121112233",
            "AB123CD", "TOYOTA", "HILUX", "BLANCO",
            "XY999ZZ", "CHEVROLET", "AVEO", "GRIS",
            "", "", "", "",
            "", "", "", ""
        ]
        for col_num, value in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col_num, value=value)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

template_service = TemplateService()
